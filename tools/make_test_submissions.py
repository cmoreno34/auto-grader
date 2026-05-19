"""
Generate synthetic test submissions from each template, with three variants:
  - _perfect.xlsx  : answer cells filled with the correct values  (should grade 100%)
  - _wrong.xlsx    : answer cells filled with a wrong option       (should grade 0%)
  - _blank.xlsx    : answer cells left empty                       (should grade 0%, status=blank)
"""

from __future__ import annotations

import json
import os
import shutil
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
TEMPLATES = Path(r"C:\Users\cesar\OneDrive - UFV\15.-UFV_ introduction to big data\final exam 2026\practice_tempaltes")
RUBRICS_DIR = ROOT / "rubrics"
OUT = ROOT / "test_submissions"


def load_rubrics():
    manifest = json.loads((RUBRICS_DIR / "manifest.json").read_text(encoding="utf-8"))
    return [json.loads((RUBRICS_DIR / f"{name}.json").read_text(encoding="utf-8")) for name in manifest["rubrics"]]


def find_template(rubric):
    hint = rubric.get("source_file_hint", "")
    if hint:
        p = TEMPLATES / hint
        if p.exists():
            return p
    # fallback: match any template that has all detect_by_sheet sheets
    for f in TEMPLATES.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            if all(s in wb.sheetnames for s in rubric["detect_by_sheet"]):
                return f
        except Exception:
            continue
    return None


def value_for_correct(q):
    """The value that, when placed in the ref cell(s), would map to the correct letter."""
    correct = q["correct"]
    t = q["type"]
    if t == "numeric_choice":
        return q["ref_sheet"], [(q["ref_cell"], _find_option_value(q, correct))]
    if t == "string_choice":
        return q["ref_sheet"], [(q["ref_cell"], _find_option_value(q, correct))]
    if t == "text_choice":
        opt = _find_option(q, correct)
        return q["ref_sheet"], [(q["ref_cell"], opt["keyword"])]
    if t == "range_choice":
        rng = next(r for r in q["ranges"] if r["letter"] == correct)
        mid = (rng["min"] + rng["max"]) / 2
        return q["ref_sheet"], [(q["ref_cell"], mid)]
    if t == "argmax_choice":
        # Set the cell matching correct.letter_if_max to a high value, others lower
        target = next(c for c in q["compare_cells"] if c["letter_if_max"] == correct)
        return q["ref_sheet"], [
            (c["cell"], 10.0 if c["cell"] == target["cell"] else 1.0)
            for c in q["compare_cells"]
        ]
    if t == "grid_row_contains":
        # Place required letters on the row that maps to the correct letter
        target_row_idx = None
        for k, v in q["row_to_letter"].items():
            if v == correct:
                target_row_idx = int(k)
                break
        if target_row_idx is None:
            return q["ref_sheet"], []
        # Parse grid_range like 'B18:F22'
        import re
        m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", q["grid_range"])
        col1, row1, col2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        # Place letters on row (row1 + target_row_idx - 1)
        target_row = row1 + target_row_idx - 1
        letters = q["required_letters"]
        # Write one letter per column starting from col1
        from openpyxl.utils import get_column_letter, column_index_from_string
        c_start = column_index_from_string(col1)
        cells = [(f"{get_column_letter(c_start + i)}{target_row}", letters[i]) for i in range(len(letters))]
        # Also fill other rows with junk letters so they don't accidentally contain required letters
        return q["ref_sheet"], cells
    if t == "string_tuple_choice":
        opt = _find_option(q, correct)
        return q["ref_sheet"], list(zip(q["ref_cells"], opt["values"]))
    if t == "freq_pair_choice":
        opt = _find_option(q, correct)
        logic = opt["logic"]
        pos_cell, neg_cell = q["ref_cells_pair"]
        if "pos>0 and neg==0" in logic:   return q["ref_sheet"], [(pos_cell, 5), (neg_cell, 0)]
        if "pos==0 and neg>0" in logic:   return q["ref_sheet"], [(pos_cell, 0), (neg_cell, 5)]
        if "pos>0 and neg>0" in logic:    return q["ref_sheet"], [(pos_cell, 3), (neg_cell, 3)]
        return q["ref_sheet"], [(pos_cell, 0), (neg_cell, 0)]
    return q.get("ref_sheet"), []


def value_for_wrong(q):
    """Pick a definitively wrong value."""
    correct = q["correct"]
    t = q["type"]
    if t == "numeric_choice":
        wrong_opt = next(o for o in q["options"] if o["letter"] != correct)
        return q["ref_sheet"], [(q["ref_cell"], wrong_opt["value"])]
    if t == "string_choice":
        wrong_opt = next(o for o in q["options"] if o["letter"] != correct)
        return q["ref_sheet"], [(q["ref_cell"], wrong_opt["value"])]
    if t == "text_choice":
        wrong_opt = next(o for o in q["options"] if o["letter"] != correct and o.get("keyword") != _find_option(q, correct).get("keyword"))
        return q["ref_sheet"], [(q["ref_cell"], wrong_opt["keyword"])]
    if t == "range_choice":
        wrong_rng = next(r for r in q["ranges"] if r["letter"] != correct)
        mid = (wrong_rng["min"] + wrong_rng["max"]) / 2
        return q["ref_sheet"], [(q["ref_cell"], mid)]
    if t == "argmax_choice":
        wrong = next(c for c in q["compare_cells"] if c["letter_if_max"] != correct)
        return q["ref_sheet"], [
            (c["cell"], 10.0 if c["cell"] == wrong["cell"] else 1.0)
            for c in q["compare_cells"]
        ]
    if t == "grid_row_contains":
        wrong_row_idx = None
        for k, v in q["row_to_letter"].items():
            if v != correct:
                wrong_row_idx = int(k); break
        import re
        m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", q["grid_range"])
        col1, row1 = m.group(1), int(m.group(2))
        target_row = row1 + (wrong_row_idx or 1) - 1
        letters = q["required_letters"]
        from openpyxl.utils import get_column_letter, column_index_from_string
        c_start = column_index_from_string(col1)
        cells = [(f"{get_column_letter(c_start + i)}{target_row}", letters[i]) for i in range(len(letters))]
        return q["ref_sheet"], cells
    if t == "string_tuple_choice":
        wrong = next(o for o in q["options"] if o["letter"] != correct)
        return q["ref_sheet"], list(zip(q["ref_cells"], wrong["values"]))
    if t == "freq_pair_choice":
        wrong = next(o for o in q["options"] if o["letter"] != correct)
        pos_cell, neg_cell = q["ref_cells_pair"]
        logic = wrong["logic"]
        if "pos>0 and neg==0" in logic:   return q["ref_sheet"], [(pos_cell, 5), (neg_cell, 0)]
        if "pos==0 and neg>0" in logic:   return q["ref_sheet"], [(pos_cell, 0), (neg_cell, 5)]
        if "pos>0 and neg>0" in logic:    return q["ref_sheet"], [(pos_cell, 3), (neg_cell, 3)]
        return q["ref_sheet"], [(pos_cell, 0), (neg_cell, 0)]
    return q.get("ref_sheet"), []


def _find_option(q, letter):
    return next(o for o in q["options"] if o["letter"] == letter)


def _find_option_value(q, letter):
    return _find_option(q, letter).get("value")


def write_cells(src: Path, dst: Path, plan: list[tuple[str, list[tuple[str, object]]]]):
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    for sheet_name, cells in plan:
        if not sheet_name or not cells:
            continue
        ws = wb[sheet_name]
        for ref, val in cells:
            ws[ref] = val
    wb.save(dst)


def main():
    OUT.mkdir(exist_ok=True)
    for rubric in load_rubrics():
        ex_id = rubric["exercise_id"]
        tpl = find_template(rubric)
        if not tpl:
            print(f"[WARN] No template for {ex_id} — skipping")
            continue
        print(f"\n{ex_id}: using template {tpl.name}")

        perfect_plan, wrong_plan = [], []
        for q in rubric["mc_questions"]:
            if q.get("type") == "manual" or "correct" not in q:
                continue  # nothing to populate for purely-conceptual questions
            ps, pcells = value_for_correct(q)
            ws_, wcells = value_for_wrong(q)
            if pcells:
                perfect_plan.append((ps, pcells))
            if wcells:
                wrong_plan.append((ws_, wcells))

        write_cells(tpl, OUT / f"{ex_id}_perfect.xlsx", perfect_plan)
        write_cells(tpl, OUT / f"{ex_id}_wrong.xlsx", wrong_plan)
        shutil.copy(tpl, OUT / f"{ex_id}_blank.xlsx")
        print(f"  wrote {ex_id}_perfect.xlsx, {ex_id}_wrong.xlsx, {ex_id}_blank.xlsx")


if __name__ == "__main__":
    main()
